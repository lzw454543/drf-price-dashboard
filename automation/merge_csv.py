"""
将永辉导出的 CSV 门店明细合并到 latest.xlsx 的永辉 sheet 中。

用法:
    python merge_csv.py <csv_path>

默认按 CSV 中出现的日期做整段替换：先删除永辉明细中同一天的旧记录，
再追加新记录，适合每日导出或按日期区间重下后修复异常数据。
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

XLSX = Path(r"C:\Users\45454\Documents\Codex\2026-08-10\new-chat-2\outputs\feishu-long-term-sync\data\latest.xlsx")
BACKUP_DIR = Path(__file__).resolve().parent / "backups"
MERGE_KEYS = ["日期", "门店编码", "商品编码"]
NUMERIC_COLS = ["销售数量", "销售金额", "促销扣款", "优惠券金额", "最终销售金额(销售金额+优惠券金额)"]
TEXT_COLS = ["商品名称", "门店名称", "品牌名称"]


def get_week_info(d: pd.Timestamp) -> tuple[str, str, int]:
    """Return month label, Monday-start week label, and in-month week days."""
    month = f"{d.month}月"
    monday = d - pd.Timedelta(days=d.weekday())
    first_day = pd.Timestamp(d.year, d.month, 1)
    first_monday = first_day - pd.Timedelta(days=first_day.weekday())
    if first_monday < first_day - pd.Timedelta(days=first_day.weekday()):
        first_monday = first_monday + pd.Timedelta(days=7)
    week_num = ((monday - first_monday).days // 7) + 1
    if week_num < 1:
        week_num = 1

    week_end = monday + pd.Timedelta(days=6)
    next_month = 1 if d.month == 12 else d.month + 1
    next_year = d.year + 1 if d.month == 12 else d.year
    month_end = pd.Timestamp(next_year, next_month, 1) - pd.Timedelta(days=1)
    overlap_start = max(monday, first_day)
    overlap_end = min(week_end, month_end)
    days_in_month = (overlap_end - overlap_start).days + 1
    return month, f"{week_num}周", days_in_month


def normalize_date(value) -> pd.Timestamp | None:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.normalize()


def load_csv_rows(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df.columns = [str(c).strip() for c in df.columns]
    df["日期"] = df["日期"].map(normalize_date)
    df = df[df["日期"].notna()].copy()

    for col in NUMERIC_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in TEXT_COLS:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["商品编码"] = pd.to_numeric(df["商品编码"], errors="coerce").astype("Int64")
    df["商品条码"] = pd.to_numeric(df["商品条码"], errors="coerce").astype("Int64")
    df["门店编码"] = df["门店编码"].astype(str).str.strip()

    rows = []
    for _, r in df.iterrows():
        d = r["日期"]
        month, week, days = get_week_info(d)
        rows.append({
            "渠道名称": "永辉",
            "月别": month,
            "周别": week,
            "当周天数": days,
            "商品编码": int(r["商品编码"]),
            "商品条码": int(r["商品条码"]),
            "商品名称": r["商品名称"],
            "门店编码": r["门店编码"],
            "门店名称": r["门店名称"],
            "品牌名称": r["品牌名称"],
            "日期": d,
            "销售数量": float(r["销售数量"]),
            "销售金额": float(r["销售金额"]),
            "促销扣款": float(r["促销扣款"]),
            "优惠券金额": float(r["优惠券金额"]),
            "最终销售金额(销售金额+优惠券金额)": float(r["最终销售金额(销售金额+优惠券金额)"]),
            "推广促销": None,
        })

    new_df = pd.DataFrame(rows)
    if new_df.empty:
        return new_df

    grouped = new_df.groupby(MERGE_KEYS, as_index=False, dropna=False)
    return grouped.agg({
        "渠道名称": "first",
        "月别": "first",
        "周别": "first",
        "当周天数": "first",
        "商品条码": "first",
        "商品名称": "first",
        "门店名称": "first",
        "品牌名称": "first",
        "销售数量": "sum",
        "销售金额": "sum",
        "促销扣款": "sum",
        "优惠券金额": "sum",
        "最终销售金额(销售金额+优惠券金额)": "sum",
        "推广促销": "first",
    }).sort_values("日期").reset_index(drop=True)


def backup_workbook() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = BACKUP_DIR / f"latest-before-yonghui-merge-{stamp}.xlsx"
    shutil.copy2(XLSX, backup)
    return backup


def replace_dates_in_workbook(new_df: pd.DataFrame) -> None:
    target_dates = {normalize_date(d) for d in new_df["日期"].dropna().unique()}
    print(f"Replacing dates in workbook: {', '.join(d.strftime('%Y-%m-%d') for d in sorted(target_dates))}")

    existing = pd.read_excel(XLSX, sheet_name=1)
    existing.columns = [str(c).strip() for c in existing.columns]
    existing["日期"] = existing["日期"].map(normalize_date)
    old_count = int(existing["日期"].isin(target_dates).sum())
    kept = existing[~existing["日期"].isin(target_dates)].copy()
    combined = pd.concat([kept, new_df], ignore_index=True, sort=False)
    combined = combined.sort_values("日期", kind="stable").reset_index(drop=True)

    headers = [str(c).strip() for c in existing.columns]
    for header in headers:
        if header not in combined.columns:
            combined[header] = None
    combined = combined[headers]

    wb = load_workbook(XLSX)
    try:
        ws = wb.worksheets[1]
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(headers)
        for row in combined.itertuples(index=False, name=None):
            values = []
            for header, value in zip(headers, row):
                if pd.isna(value):
                    value = None
                elif header == "日期":
                    value = pd.Timestamp(value).to_pydatetime()
                values.append(value)
            ws.append(values)
        wb.save(XLSX)
    finally:
        wb.close()

    print(f"  Removed {old_count} old rows and wrote {len(combined)} total rows to sheet 1")


def print_summary(new_df: pd.DataFrame | None = None) -> None:
    if new_df is not None:
        print("\nNew data summary:")
        source = new_df.copy()
    else:
        print("\nWorkbook verification:")
        source = pd.read_excel(XLSX, sheet_name=1)
        source.columns = [str(c).strip() for c in source.columns]
        source["日期"] = source["日期"].map(normalize_date)
        source = source[source["日期"].notna()].copy()

    if new_df is None:
        dates = [pd.Timestamp("2026-08-17"), pd.Timestamp("2026-08-18"), pd.Timestamp("2026-08-19")]
        source = source[source["日期"].isin(dates)]

    summary = (
        source.groupby("日期")
        .agg(rows=("商品编码", "size"), qty=("销售数量", "sum"), sales=("最终销售金额(销售金额+优惠券金额)", "sum"), stores=("门店编码", "nunique"))
        .reset_index()
    )
    for _, row in summary.iterrows():
        print(f"  {row['日期'].strftime('%Y-%m-%d')}: rows={int(row['rows'])}, qty={row['qty']:.0f}, sales={row['sales']:.2f}, stores={int(row['stores'])}")


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python merge_csv.py <csv_path>")
    csv_path = Path(sys.argv[1])
    if not csv_path.exists():
        raise SystemExit(f"CSV not found: {csv_path}")
    if not XLSX.exists():
        raise SystemExit(f"Workbook not found: {XLSX}")

    print(f"Reading CSV: {csv_path}")
    new_df = load_csv_rows(csv_path)
    if new_df.empty:
        print("No valid rows found.")
        return
    print_summary(new_df)

    backup = backup_workbook()
    print(f"\nBackup created: {backup}")
    replace_dates_in_workbook(new_df)
    print_summary()


if __name__ == "__main__":
    main()
